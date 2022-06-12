print('file 100')
import openpyxl
wb=openpyxl.load_workbook("C:\\Users\\Home\\Desktop\\New folder\\empty_book.xlsx")
sheets=wb.sheetnames
print(wb.active.title)
print('file 101')
sh1=wb['sheet1']
data=sh1['B1'].value
print(data)
print('file 102')
sh1.cell(row=10,coloum=10,value='my')
wb.save("C:\\Users\\Home\\Desktop\\New folder\\empty_book.xlsx")

